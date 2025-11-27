from django.shortcuts import render
from apps.student.models import Student
from django.http import JsonResponse
import json
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import os
from django.conf import settings
import uuid
import hashlib
import traceback
from openpyxl import load_workbook,Workbook  # type: ignore


# Create your views here.
def get_students(request):
    try :
        obj_students = Student.objects.all().values()
        students = list(obj_students)
        return  JsonResponse({'code': 1, 'data':students, 'safe':False})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg':"获取学生消息异常，具体错误：" + str(e)})


@csrf_exempt  # ✅ 禁用 CSRF 验证
@require_http_methods(["POST"])
def query_students(request):
    data = json.loads(request.body.decode('utf-8'))
    try :
        input_str = data.get('keyword', '').strip()
        
        if not input_str:
            # 空查询返回所有学生
            students = list(Student.objects.all().values())
            return JsonResponse({'code': 1, 'data': students})
        
        # 构建查询条件
        query = Q()
        search_fields = ['sno__icontains', 'name__icontains', 'gender__icontains', 
                        'mobile__icontains', 'email__icontains', 'address__icontains',
                        'birthday__icontains']
        
        for field in search_fields:
            query |= Q(**{field: input_str})
        
        students = list(Student.objects.filter(query).values())
        return  JsonResponse({'code': 1, 'data':students})
    except json.JSONDecodeError:
        return JsonResponse({'code': 0, 'msg': '请求数据格式错误'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg':"查询学生消息异常，具体错误：" + str(e)})


@csrf_exempt  # ✅ 禁用此视图的 CSRF 验证
@require_http_methods(["POST"])
def is_exists_sno(request):
    data = json.loads(request.body.decode('utf-8'))
    try:
        obj_student = Student.objects.filter(sno=data['sno'])
        if obj_student.count() == 0:
            return JsonResponse({'code':1, 'exists':False})
        else:
            return JsonResponse({'code':1, 'exists':True})
    except Exception as e:
        return JsonResponse({'code':0, 'msg':"校验学好失败，原因：" + str(e)})


@csrf_exempt  # ✅ 禁用此视图的 CSRF 验证
@require_http_methods(["POST"])    
def add_students(request):
    data = json.loads(request.body.decode('utf-8'))
    try:
        obj_student = Student(sno=data['sno'],name=data['name'],gender=data['gender'],birthday=data['birthday'],mobile=data['mobile'],email=data['email'],address=data['address'],image=data['image'])
        obj_student.save()
        obj_students = Student.objects.all.values()
        students = list(obj_students)
        return JsonResponse({'code': 1,'msg': '添加成功', 'data': students})
    except Exception as e:
        return JsonResponse({'code':0, 'msg':"添加数据失败，原因" + str(e)})


@csrf_exempt  # ✅ 禁用此视图的 CSRF 验证
@require_http_methods(["POST"])    
def update_students(request):
    data = json.loads(request.body.decode('utf-8'))
    try:
        obj_student = Student.objects.get(sno=data['sno'])
        obj_student.name = data['name']
        obj_student.gender = data['gender']
        obj_student.birthday = data['birthday']
        obj_student.mobile = data['mobile']
        obj_student.email = data['email']
        obj_student.address = data['address']
        obj_student.image = data['image']
        obj_student.save()
        obj_students = Student.objects.all().values()
        students = list(obj_students)
        return JsonResponse({'code': 1,'msg': '添加成功', 'data': students})
    except Exception as e:
        return JsonResponse({'code':0, 'msg':"修改数据失败，原因" + str(e)})


@csrf_exempt  # ✅ 禁用此视图的 CSRF 验证
@require_http_methods(["POST"])    
def delete_students(request):
    data = json.loads(request.body.decode('utf-8'))
    try:
        obj_student = Student.objects.get(sno=data['sno'])
        obj_student.delete()
        obj_students = Student.objects.all.values()
        students = list(obj_students)
        return JsonResponse({'code': 1,'msg': '删除成功', 'data': students})
    except Exception as e:
        return JsonResponse({'code':0, 'msg':"删除数据失败，原因" + str(e)})


@csrf_exempt  # ✅ 禁用此视图的 CSRF 验证
@require_http_methods(["POST"])    
def delete_all_students(request):
    data = json.loads(request.body.decode('utf-8'))
    print('【data】', type(data), data)
    try:
        for one_student in data['student']:
            obj_student = Student.objects.get(sno=one_student['sno'])
            obj_student.delete()

        obj_students = Student.objects.all().values()
        students = list(obj_students)
        return JsonResponse({'code': 1,'msg': '批量删除成功', 'data': students})
    except Exception as e:
        return JsonResponse({'code':0, 'msg':"批量删除数据失败，原因" + str(e)})

@csrf_exempt 
@require_http_methods(["POST"]) 
def import_students_excel(request):
    rev_file = request.FILES.get('excel')
    if not rev_file:
        return JsonResponse({'code':0,'msg':'Excel文件不存在！'})
    new_name = get_random_str()
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1])
    try:
        f = open(file_path, 'wb')
        for i in rev_file.chunks():
            f.write(i)
        f.close()
    except Exception as e:
        return JsonResponse({'code':0,'msg':str(e)})
    
    ex_students = read_excel_dict(file_path)
    success = 0
    error = 0
    error_snos = []
    for one_student in ex_students:
        try:
            obj_student = Student(sno=one_student['sno'],name=one_student['name'],gender=one_student['gender'],birthday=one_student['birthday'],mobile=one_student['mobile'],email=one_student['email'],address=one_student['address'])
            obj_student.save()
            success += 1
        except:
            error += 1
            error_snos.append(one_student.sno)
    
    obj_student = Student.objects.all().values()
    students = list(obj_student)
    return JsonResponse({'code':1,'success':success,'error':error,'errors':error_snos,'data':students})

def get_random_str():
    uuid_val = uuid.uuid4()
    uuid_str = str(uuid_val).encode('utf-8')
    md5 = hashlib.md5()
    md5.update(uuid_str)
    return md5.hexdigest()

def read_excel_dict(path:str):
    #实例化一个workbook
    workbook = load_workbook(path)
    #实例化一个sheet
    sheet = workbook['student']
    #定义一个变量存储最终的数据
    students = []
    #准备key
    keys = ['sno','name','gender','birthday','mobile','email','address']
    #遍历
    for row in sheet.rows:
        temp_dict = {}
        for index,cell in enumerate(row):
            temp_dict[keys[index]] = cell.value
        students.append(temp_dict)
    return students

@csrf_exempt 
@require_http_methods(["POST"]) 
def upload(request):
    rev_file = request.FILES.get('avatar')
    if not rev_file:
        return JsonResponse({'code':0, 'msg':'图片不存在！'})
    new_name = get_random_str()
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1])
    try:
        f = open(file_path,'wb')
        for i in rev_file.chunks():
            f.write(i)
        return JsonResponse({'code':1,'name':new_name + os.path.splitext(rev_file.name)[1]})

    except Exception as e:
        return JsonResponse({'code':0,'msg':traceback.format_exc()})
    
@csrf_exempt 
@require_http_methods(["POST"]) 
def export_student_excel(request):
    obj_students = Student.objects.all().values()
    students = list(obj_students)
    excel_name = get_random_str() + ".xlsx"
    path = os.path.join(settings.MEDIA_ROOT,excel_name)
    write_to_excel(students,path)
    return JsonResponse({'code':1, 'name':excel_name})

def write_to_excel(data:list,path:str):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'student'
    if not data:
        return
    # 表头用第一条字典的 key
    headers = list(data[0].keys())
    sheet.append(headers)                       # 写表头
    for row in data:
        sheet.append([row.get(k, '') for k in headers])  # 写数据行
    workbook.save(path)